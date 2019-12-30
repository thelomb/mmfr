from openpyxl import load_workbook
import sys


class Reader():
    def __init__(self):
        self.header = []
        self.range = []

    def parse(self, path, sheetname):
        if sheetname is None:
            sheetname = 'Sheet1'
        wb = load_workbook(path, read_only=True)
        sheet = wb[sheetname]
        for row in sheet.iter_rows(max_row=1, values_only=True):
            self.header = list(row)
        # for i in range(1, sheet.max_row):
        #     line = dict()
        #     x = 0
        #     for header in list(sheet.iter_rows())[0]:
        #         line[header.value] = list(sheet.iter_rows())[i][x].value
        #         x += 1
        #     self.range.append(line)
        #     print('processed row:', i)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            self.range.append(list(row))

        return (self.header, self.range)


class XLType():
    def __init__(self, cls):
        self.cls = cls

    def clean(self, v):
        return getattr(sys.modules[__name__], self.cls)().clean(v)


class XLString():
    @staticmethod
    def clean(v):
        if v:
            return str(v).strip()
        return v


class XLDate():
    @staticmethod
    def clean(v):
        if v:
            return v.date().isoformat()
        return v


class XLBool:
    @staticmethod
    def clean(v):
        return v in ['Yes', 'YES']


class XLEnum:
    mapping = {}
    default = None

    @classmethod
    def clean(cls, value):
        if isinstance(value, str):
            space = str.maketrans(dict.fromkeys(' '))
            value = value.translate(space)
            return cls.mapping[value.lower()]
        if cls.default:
            return cls.default
        return value


class MMFMasterFeedType(XLEnum):
    mapping = {'fder': 'FDER',
               'mstr': 'MSTR',
               'none': 'NONE',
               'feeder': 'FDER',
               'master': 'MSTR'
               }


class MMFType(XLEnum):
    mapping = {'shorttermlvnavmoneymarketfund': 'STLV',
               'shorttermpublicdebtcnavmoneymarketfund': 'STCN',
               'shorttermvnavmoneymarketfund': 'STVN',
               'standardvnavmoneymarketfund': 'SDVN',
               'ShortTermLVNAVMoneyMarketFund': 'STLV',
               'ShortTermPublicDebtCNAVMoneyMarketFund': 'STCN',
               'ShortTermVNAVMoneyMarketFund': 'STVN',
               'StandardVNAVMoneyMarketFund': 'SDVN',
               'shorttermlvnavmmf': 'STLV',
               'shorttermpublicdebtcnavmmf': 'STCN',
               'shorttermvnavmmf': 'STVN',
               'standardvnavmmf': 'SDVN'
               }


class MMFLegalFramework(XLEnum):
    mapping = {'UCITS': 'UCIT',
               'AIFD': 'AIFD',
               'UCIT': 'UCIT',
               'ucit': 'UCIT',
               'aifd': 'AIFD',
               'UndertakingsForCollectiveInvestmentInTransferableSecurities': 'UCIT',
               'AlternativeInvestmentFund': 'AIFD'
               }


class XLList:
    @staticmethod
    def clean(v):
        return v.split(';')


class XLNumeric:
    @staticmethod
    def clean(v):
        if isinstance(v, float):
            return max(0, round(v, 5))
        else:
            return v


class PositionPartySector(XLEnum):
    mapping = {
        'SubjectToRegulationSupranationalPublicBody': 'SRSN',
        'SubjectToRegulationSovereign': 'SRSB',
        'SubjectToRegulationPublicBody': 'SRPB',
        'SubjectToRegulationCentralBank': 'SRCB',
        'Regional': 'RGNL',
        'OtherFinancialCorporation': 'OFCP',
        'NotSubjectToRegulationSupranationalPublicBody': 'NRSN',
        'NotSubjectToRegulationSovereign': 'NRSB',
        'NotSubjectToRegulationPublicBody': 'NRPB',
        'NotSubjectToRegulationCentralBank': 'NRCB',
        'NonFinancialCorporation': 'NFIN',
        'NationalPublicBody': 'NTPB',
        'Local': 'LOCA',
        'CreditInstitution': 'CDTI'
    }
    default = 'OFCP'


class ValuationType(XLEnum):
    mapping = {'MarkToModel': 'MTMO',
               'marktomodel': 'MTMO',
               'MarkToMarket': 'MTMA',
               'marktomarket': 'MTMA',
               'AmortisedCost': 'AMCS'
               }


class AssessmentResultType(XLEnum):
    mapping = {
        'Favourable': 'FVRB',
        'NotApplicable': 'NOAP',
        'Unfavourable': 'UFVB',
        'NotPerformed': 'NOVF'
    }
    default = 'NOVF'


class AssetType(XLEnum):
    mapping = {
        'SimpleTransparentStandardisedAssetBackedCommercialPaper': 'STSA',
        'Securitisation': 'SCRT',
        'AssetBackedCommercialPaper': 'ABCP',
        'assetbackedcommercialpaper': 'ABCP',
        'SimpleTransparentStandardisedSecuritisation': 'STSS',
        'MoneyMarketInstrument': 'MMII',
        'moneymarketinstrument': 'MMII',
        'FinancialDerivativeInstrumentOverTheCounter': 'OTCD',
        'FinancialDerivativeInstrumentRegulatedMarket': 'RMTD',
        'UnitOrShareOfOtherMoneyMarketFund': 'MMFT',
        'unitorshareofothermoneymarketfund': 'MMFT',
        'DepositsWithCreditInstitution': 'DPSC',
        'depositswithcreditinstitution': 'DPSC',
        'AncillaryLiquidAsset': 'ANLA',
        'ancillaryliquidasset': 'ANLA',
        'ReverseRepurchaseAgreement': 'RVPO',
        'RepurchaseAgreement': 'REPO'
    }


class ContractType(XLEnum):
    mapping = {
        'ContractForDifference': 'CFDS',
        'ForwardRateAgreement': 'FRAS',
        'Futures': 'FUTR',
        'Forward': 'FORW',
        'Option': 'OPTN',
        'Swap': 'SWAP',
        'Swaption': 'SWPT',
        'Other': 'OTHR',
        'ForwardsOnASwap': 'FWOS',
        'FuturesOnSwap': 'FONS'
    }


class UnderlyingType(XLEnum):
    mapping = {
        'Currency': 'CURR',
        'IndexOfCurrencies': 'CIDX',
        'IndexOfInterestRates': 'IIDX',
        'Interest': 'INTR',
        'InterestRateAndCurrency': 'INCU'
    }


class FinancialUnderlyingType(XLEnum):
    mapping = {
        'CommercialMortgage': 'CMBS',
        'ConsumerLoans': 'CSML',
        'CreditCardReceivables': 'CCRB',
        'Leasing': 'LESG',
        'LoansToCorporatesOrSME': 'LTCS',
        'OtherAsset': 'OTHR',
        'otherasset': 'OTHR',
        'ResidentialMortgage': 'RMBS',
        'TradeReceivables': 'TDRB',
        '': 'NONE'
    }


class PerformanceRange(XLEnum):
    mapping = {
        'twotosevendays': 'D2T7',
        'eighttotwentyninedays': 'D829',
        'lessoneday': 'LS1D',
        'abovethirtydays': 'A30D'

               }

class Frequency(XLEnum):
    mapping = {
        'month': 'MNTH',
        'twiceamonth': 'TWMN',
        'weekly': 'WEEK',
        'daily': 'DAIL',
        'adhoc': 'ADHO',
        'none': 'NONE'
    }
    default = 'NONE'

class Arrangement(XLEnum):
    mapping = {
        'SuspensionOfRedemption': 'SPRN',
        'Gates': 'GATE',
        'LiquidityFeesOnRedemptions': 'RDLF',
        'OtherArrangement': 'OTHR'
    }

class XLNone:
    @staticmethod
    def clean(v):
        return v

class NetAssetBasis(XLEnum):
    mapping = {
        'cnav': 'CNAV',
        'navl': 'NAVL'
    }