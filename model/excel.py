from openpyxl import Workbook, load_workbook

class Reader():
    def __init__(self, sheetname='Sheet1'):
        self.data = {}
        self.range = []
        self.sheetname = sheetname

    def parse(self, path):
        wb = load_workbook(path, read_only=True)
        sheet = wb[self.sheetname]
        for i in range(1, sheet.max_row):
            line = dict()
            x = 0
            for header in list(sheet.iter_rows())[0]:
                line[header.value] = list(sheet.iter_rows())[i][x].value
                x += 1
            self.range.append(line)
        return self.range




